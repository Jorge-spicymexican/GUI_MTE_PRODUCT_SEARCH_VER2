# Licensed to TCI, LLC
# Located in W132N10611 Germantown, Wisconsin
# This file is owned and copyrighted by TCI, LLC
#  Author: Jorge Jesus Jurado-Garcia
#  Title: Product Specialist Intern
#   Project Description: Marketing department MTE Product selection
#   Goal: create an easy to use GUI where the marketing department can
#          use for any price increases. Anyone with no background in CCS
#
#  Date of Creation: 12/29/2021
#  Rev:

import tkinter as tk
from tkinter import ttk

from selenium import webdriver
from openpyxl import Workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import file_lookup_imports as FM
import noti_and_checker as noti


def selenium(array_product_name, web_driver, target, date, author, directory):
    # main website of MTE Corporation
    main_page = "https://www.mtecorp.com/click-find/"
    driver = webdriver.Chrome(executable_path=web_driver)
    # Goes to main page
    driver.get(main_page)
    # to maximize the browser window
    driver.minimize_window()
    driver = go(target, driver)
    wb = scrape(array_product_name, driver)
    save_excel(wb, date, author, directory)


def go(selection, driver):
    if selection == "RL":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[1]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "RLW":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[2]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
        # input_product_name(array_Product_Name)
    elif selection == "DVS":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[3]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "SWG":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[5]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "SWN":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[6]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "SWGM":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[7]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "MAP":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[8]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "MAEP":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[9]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "DVT":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[3]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "RF3":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[12]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    return driver


def scrape(array_product_name, driver):
    # these two arrays will be used to store the values havest from the webpage
    description = []
    unit_price = []
    #  switches where selenium looks at by going to the second screen
    #  function below is just so the bot can focus on new opened windows handler
    driver.switch_to.window(driver.window_handles[1])

    # for loop for the size of the array named this look will clear search bar, enter excel product name,
    # press enter, locate the next button, and click on the button for the next page.
    for ll in range(len(array_product_name)):
        # clears search bar
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".plxsty_pid"))).clear()
        # inputs text string into the search bar and waits to execute for 20 seconds
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".plxsty_pid"))).send_keys(
            array_product_name[ll])
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".plxsty_pid"))).send_keys(
            Keys.ENTER)
        # looks for element by full xpath and clicks with arguements[0] is are fullfilled.
        accept_bar = driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div[3]/table/tbody/tr/td[1]")
        driver.execute_script("arguments[0].click();", accept_bar)
        d1 = driver.find_element(By.XPATH, "/html/body/div/table[4]/tbody/tr[2]/td/table/tbody/tr[2]/td[3]")
        d1 = d1.text
        description.append(d1)
        p1 = driver.find_element(By.XPATH, "/html/body/div/table[4]/tbody/tr[2]/td/table/tbody/tr[2]/td[4]")
        p1 = p1.text
        unit_price.append(p1)
        # This line excute_script is to move back to the last opened page.
        driver.execute_script("window.history.go(-1)")

    # close the web browser and finish
    driver.close()
    wb = new_excel(array_product_name, description, unit_price)
    return wb


def new_excel(array_product_name, description, unit_price):
    # create a Workbook object.
    work_book = Workbook()
    sh = work_book.active
    sh.title = "Sheet1"
    for j in range(0, len(array_product_name)):
        sh.cell(row=j + 1, column=1).value = array_product_name[j]
        sh.cell(row=j + 1, column=2).value = description[j]
        sh.cell(row=j + 1, column=3).value = unit_price[j]
    return work_book


def save_excel(wb, date, author, directory):
    # create a file name using the date, and author
    # check if date has any / or \ -
    if directory != "":
        file = directory + "/" + author + "_" + date + "_" + "Result.xlsx"
    else:
        file = author + "_" + date + "_" + "Result.xlsx"
    wb.save(filename=file)
    showinfo(
        title='New Excel Location : ',
        message=file
    )

"""
Function for Creating new window for new settings and configurations
"""
def New_Product_Window():
    # creates a Tk() object
    master = tk.Tk()

    master.title('TCI LLC - MTE Product Search')
    master.resizable(False, False)
    # ensure that a window is always at the top of the stacking order
    master.attributes('-topmost', 1)
    master.geometry("250x230")

    master.iconbitmap('./assets/tci_logo_Csx_icon.ico')
    master.grid()


"""
This class created the main window for UI and Setups of buttons and entry points. 
It also assigns functions and actions to each specific functions when clicked.
"""
class MainFrame(ttk.Frame):
    run_excel: bool

    # Initialization
    def __init__(self, container):
        super().__init__(container)
        # field options
        options = {'padx': 5, 'pady': 5}

        # Name label
        ttk.Label(self, text="Name :").grid(column=0, row=0, sticky=tk.W, **options)

        # Name entry
        self.projectname = tk.StringVar()
        ttk.Entry(self, textvariable=self.projectname).grid(column=1, columnspan=2, row=0, sticky=tk.EW, **options)

        # file_name
        self.filename = str()

        # web driver location
        self.web_driver = str()

        # file_name
        self.new_excel_path = str()

        # Dump excel here
        ttk.Button(self, text="New Location",
                   command=self.new_excel_location).grid(row=1, column=0, sticky=tk.EW, **options)

        # web driver button
        ttk.Button(self, text='Web Driver',
                   command=self.select_webdriver).grid(column=1, columnspan=2, row=1, sticky=tk.EW, **options)

        # excel button
        ttk.Button(self, text='Insert Raw Data',
                   command=self.select_excel).grid(column=0, row=2, columnspan=3, sticky=tk.EW, **options)

        self.run_excel = False

        # Create A checkbox Button
        self.btn = tk.Checkbutton(self, text='Use Imported Data', command=self.checkbox,
                                  variable=self.run_excel)
        # self.btn.configure(bg='#ffb3fe')
        self.btn.grid(row=3, column=0, columnspan=3, sticky=tk.EW)

        # start button
        self.button = ttk.Button(self, text="Start Scrap", command=self.start)
        self.button.grid(row=4, column=0, columnspan=3, sticky=tk.EW, **options)

        # Close Button
        tk.Button(self, text="Exit", command=self.close,
                  bg="Red", fg="White").grid(row=5, column=0, columnspan=3, sticky=tk.EW, **options)

        # add padding to the frame and show it
        self.grid(padx=10, pady=10, sticky=tk.NSEW)

    def new_excel_location(self):
        self.new_excel_path = FM.folder_lookup()

    def start(self):
        """
        In this function we will first check that the excel can be read
        afterwards checking if the excel can be read we will then check the chrome driver being used
        """
        if self.run_excel is True:
            if self.filename == "":
                "Tell the user that they have not loaded an excel"
                noti.InsertExcel()
                return
            else:
                "Check if the excel can be read if so save the values and store the data into an array " \
                "afterwards run a different that checks the entries being inserted."
                self.Product_array = FM.load_excel(self.filename)
                if self.Product_array is None:
                    return
        else:
            """
            If wer are not reading from a raw excel allow the user to select which MTE Product Family to Read from
            and store the result to a a varibale.
            """
            New_Product_Window()

    @property
    def check_entry(self):
        check = True
        # Retrieve the of name, date, and MTE Selection
        inputs = [self.projectname.get(), self.selected_product.get(), self.filename,
                  self.web_driver]
        if inputs[0] == '':
            showerror(
                title='Error-Name',
                message='Please type in name.'
            )
            check = False
        if inputs[1] == '':
            showerror(
                title='Error-Selection',
                message='User did not Selected Product Line, Please Check.'
            )
            check = False
        if inputs[2] == '':
            showerror(
                title='Error-Excel File',
                message='User did not Inserted an Excel File.'
            )
            check = False
        if inputs[3] == '':
            showerror(
                title='Error-Webdriver',
                message='User did not Inserted a Selenium based Webdriver.'
            )
            check = False
        if check:
            showinfo(
                title='Settings',
                message='Settings are Configured'
            )
            self.show_selected_product()
            product_name = load_excel(inputs[3])
            if len(product_name) != 0:
                if self.web_driver[-16:] == "chromedriver.exe":
                    selenium(product_name, inputs[4], inputs[2], inputs[1], inputs[0], self.new_excel_path)
                else:
                    showerror(
                        title='Error-Driver',
                        message='Does not support this Driver. Make sure your Driver is named like this...chromedriver'
                    )

    def select_excel(self):
        self.filename = FM.file_lookup()

    def select_webdriver(self):
        self.web_driver = FM.file_lookup()

    def close(self):
        self.quit()

    def checkbox(self):
        if self.run_excel is False:
            self.run_excel = True
        else:
            self.run_excel = False


"""
Main Object Creation - this class setups the title, Window Attributes,
 Width and height, and Logo for UI
"""


class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title('TCI LLC - MTE Product Search')
        self.resizable(False, False)
        # ensure that a window is always at the top of the stacking order
        self.attributes('-topmost', 1)

        window_width = 250
        window_height = 220

        # get screen dimension
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()

        # find the center point
        center_x = int(screen_width / 2 - window_width / 2)
        center_y = int(screen_height / 2 - window_height / 2)

        # create the screen on window console
        self.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')

        # changing the Tinker Logo into the TCI logo instead for our development
        self.iconbitmap('./assets/tci_logo_Csx_icon.ico')
        frm = ttk.Frame(self, padding=1)
        frm.grid()


if __name__ == "__main__":
    app = App()
    MainFrame(app)
    app.mainloop()
