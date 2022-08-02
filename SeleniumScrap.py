# Licensed to TCI, LLC
# Located in W132N10611 Germantown, Wisconsin
# This file is owned and copyrighted by TCI, LLC
#  Author: Jorge Jesus Jurado-Garcia
#  Title: Product Specialist Intern
#   Project Description: Marketing department MTE Product selection
#   Goal: create an easy to use GUI where the marketing department can
#          use for any price increases. Anyone with no background in CCS
#
#  Date of Creation: 12/29/2021
#  Rev:

from selenium import webdriver
from openpyxl import Workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import file_lookup_imports as FM
import noti_and_checker as noti

def selenium(array_product_name, web_driver, author, directory):
    # main website of MTE Corporation
    main_page = "https://www.mtecorp.com/click-find/"
    driver = webdriver.Chrome(executable_path=web_driver)
    # Goes to main page
    driver.get(main_page)
    # to maximize the browser window
    driver.maximize_window()

    for i in array_product_name:
        print(array_product_name[i])

    if len(array_product_name) !=0:
        #get the first element of the array
        print("PASSED VALUE:", array_product_name[1] )
        #KEY_VALUE = GetProductType(array_product_name[0])

    print(KEY_VALUE)

    driver = go(KEY_VALUE, driver)

    # SCRAPE DATE FOR THIS WEBPAGE
     #wb = scrape(array_product_name, driver)

    # SAVE EXCEL FILE
    # save_excel(wb, date, author, directory)

def TestChromeDriver(web_driver):
    # main website of MTE Corporation
    main_page = "https://www.mtecorp.com/click-find/"
    try:
        driver = webdriver.Chrome(executable_path=web_driver)
        driver.maximize_window()
        driver.get(main_page)
        driver.close()
        return True
    except Exception as e:
        print(e)
        # Tell the user what reversion of Code is being Used and where to find there web-browser type
        paths = [r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                 r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"]
        version = list(filter(None, [noti.get_version_via_com(p) for p in paths]))[0]

        # Tell the user to download the this version Selenium Webdriver from website url
        Selenium_ChromeDownload_URL = "https://chromedriver.chromium.org/downloads"
        # send notification of issue found
        FM.WebDriverVersionIssue(version)
        driver.close()
        return False

def GetProductType(NME):
    print("GET PRODUCT TYPE ARGUEMENTS:", NME)
    # check if value matches with these first four strings
    if "MAEP" in NME[0:4]:
        ProductInfo = "MAEP"
    elif "SWGM" in NME[0:4]:
        ProductInfo = "SWGM"
    elif "RF3" in NME[0:3]:
        ProductInfo = "RF3"
    elif "MAP" in NME[0:3]:
        ProductInfo = "MAP"
    elif "SWN" in NME[0:3]:
        ProductInfo = "SWN"
    elif "SWG" in NME[0:3]:
        ProductInfo = "SWG"
    elif "DVS" in NME[0:3]:
        ProductInfo = "DVS"
    elif "DVT" in NME[0:3]:
        ProductInfo = "DVT"
    elif "RLW" in NME[0:3]:
        ProductInfo = "RLW"
    elif "RL" in NME[0:2]:
        ProductInfo = "RL"
    else:
        ProductInfo = "ERROR"

    return ProductInfo


def go(selection, driver):
    if selection == "RL":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[1]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "RLW":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[2]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
        # input_product_name(array_Product_Name)
    elif selection == "DVS":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[3]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "SWG":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[5]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "SWN":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[6]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "SWGM":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[7]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "MAP":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[8]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "MAEP":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[9]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "DVT":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[3]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "RF3":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[12]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    return driver


def scrape(array_product_name, driver):
    # these two arrays will be used to store the values havest from the webpage
    description = []
    unit_price = []
    #  switches where selenium looks at by going to the second screen
    #  function below is just so the bot can focus on new opened windows handler
    driver.switch_to.window(driver.window_handles[1])

    # for loop for the size of the array named this look will clear search bar, enter excel product name,
    # press enter, locate the next button, and click on the button for the next page.
    for ll in range(len(array_product_name)):

        # add a try catch method for getting the data for each Product Automated Search
        # clears search bar
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".plxsty_pid"))).clear()
        # inputs text string into the search bar and waits to execute for 20 seconds
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".plxsty_pid"))).send_keys(
            array_product_name[ll])
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".plxsty_pid"))).send_keys(
            Keys.ENTER)


        # looks for element by full xpath and clicks with arguements[0] is are fullfilled.
        accept_bar = driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div[3]/table/tbody/tr/td[1]")
        driver.execute_script("arguments[0].click();", accept_bar)
        d1 = driver.find_element(By.XPATH, "/html/body/div/table[4]/tbody/tr[2]/td/table/tbody/tr[2]/td[3]")
        d1 = d1.text
        description.append(d1)
        p1 = driver.find_element(By.XPATH, "/html/body/div/table[4]/tbody/tr[2]/td/table/tbody/tr[2]/td[4]")
        p1 = p1.text
        unit_price.append(p1)
        # This line excute_script is to move back to the last opened page.
        driver.execute_script("window.history.go(-1)")

    # close the web browser and finish
    driver.close()

    wb = new_excel(array_product_name, description, unit_price)
    return wb


def new_excel(array_product_name, description, unit_price):
    # create a Workbook object.
    work_book = Workbook()
    sh = work_book.active
    sh.title = "Sheet1"
    for j in range(0, len(array_product_name)):
        sh.cell(row=j + 1, column=1).value = array_product_name[j]
        sh.cell(row=j + 1, column=2).value = description[j]
        sh.cell(row=j + 1, column=3).value = unit_price[j]
    return work_book


def save_excel(wb, date, author, directory):
    # create a file name using the date, and author
    # check if date has any / or \ -
    if directory != "":
        file = directory + "/" + author + "_" + date + "_" + "Result.xlsx"
    else:
        file = author + "_" + date + "_" + "Result.xlsx"
    wb.save(filename=file)
    """showinfo(
        title='New Excel Location : ',
        message=file
    )"""
