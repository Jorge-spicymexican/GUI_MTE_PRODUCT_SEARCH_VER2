from tkinter.messagebox import showinfo

import pandas.core.frame

import noti_and_checker as noti
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException, ReadOnlyWorkbookException
import pandas as pd
from tkinter import filedialog as fd

"""
This functions allows the User to Selected any type of files to be imported for UI Interface. It will then
display the file imported and show the user where the was imported.
"""


def file_lookup():
    filetypes = (
        ('All files', '*.*'),
        ('CSV files', '*.csv'),
        ('Excel files', '*.xlsx')
    )
    filename = fd.askopenfilename(
        title='Insert File',
        initialdir='/',
        filetypes=filetypes)
    if filename is None:
        noti.FileInsertErrorNotfi()
    else:
        noti.FileInsertedNotfi(filename)
    return filename


"""
Folder lookup
"""


def folder_lookup():
    directory = fd.askdirectory(
    )
    if directory != "":
        showinfo(
            title='Selected Directory',
            message=directory
        )
    return directory


"""
This function tries to read the excel with 3 different functions. When it can read 
the excel file it will will just end up notifying the user and give the user ample time to fix this issue. This is 
in order to make sure the user can get his settings configured before starting program
"""


def load_excel(filepath):
    df = None
    if filepath:
        try:
            filepath = r"{}".format(filepath)
            df = pd.read_csv(filepath, index_col=None, na_values=['NA'], usecols="A")
        except ValueError:
            try:
                df = pd.read_excel(filepath, index_col=None, na_values=['NA'], usecols="A")
            except ValueError:
                noti.ExcelReadError()
            except FileNotFoundError:
                noti.NotFoundError()
            except ImportError:
                noti.ImportError()
        except FileNotFoundError:
            try:
                df = pd.read_excel(filepath, index_col=None, na_values=['NA'], usecols="A")
            except ValueError:
                noti.ExcelReadError()
            except FileNotFoundError:
                noti.NotFoundError()
            except ImportError:
                noti.ImportError()
        except ImportError:
            try:
                df = pd.read_excel(filepath, index_col=None, na_values=['NA'], usecols="A")
            except ValueError:
                noti.ExcelReadError()
            except FileNotFoundError:
                noti.NotFoundError()
            except ImportError:
                noti.ImportError()
    if df is None:
        try:
            df = load_workbook(filepath)
        except InvalidFileException:
            noti.ExcelReadError()

        except ReadOnlyWorkbookException:
            noti.ExcelReadonlyError()

    """
    Check what is the data type of these sheet and Insert the cells of this excel sheet into an array.
    """
    print(type(df))
    if type(df) is not pandas.core.frame.DataFrame:
        if 'Sheet1' in df.sheetnames:
            sh = df["Sheet1"]
            row_ct = sh.max_row
            array_Product_Name = []
            for i in range(1, row_ct):
                array_Product_Name.append(sh.cell(row=i, column=1).value)
                # Will have to know how to grab information from excel and such
            return array_Product_Name
    else:
        return df
